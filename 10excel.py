import openpyxl
from tabulate import tabulate

"""
Enseñar esto de primero
excel_dataframe = openpyxl.load_workbook("C:\\Users\\luisc\\desktop\\personas.xlsx")

dataframe =excel_dataframe.active

data = []

for row in range(1, dataframe.max_row):
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column):
        print(col[row].value)"""

"""
LUEGO ENSEÑAR ESTO
excel_dataframe = openpyxl.load_workbook("C:\\Users\\luisc\\desktop\\personas.xlsx")

dataframe =excel_dataframe.active

data = []

for row in range(1, dataframe.max_row):
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column):
        _row.append(col[row].value)
    data.append(_row)

print(tabulate(data))#primero hacerlo solo con data y luego con tabulate"""



excel_dataframe = openpyxl.load_workbook("C:\\Users\\luisc\\desktop\\personas.xlsx")

dataframe =excel_dataframe.active

data = []

for row in range(1, dataframe.max_row):
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column):
        _row.append(col[row].value)
    data.append(_row)

headers = ["#","Id","Nombre","Compañia","Email","MAC Address"]
headers_align = (("center",)*6)
print(tabulate(data, headers=headers,tablefmt="fancy_grid", colalign=headers_align))